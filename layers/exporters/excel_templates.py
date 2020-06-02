import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
try:
    from exporters.matrix_gen import MatrixGen
except ModuleNotFoundError:
    from ..exporters.matrix_gen import MatrixGen


class BadTemplateException(Exception):
    pass

class ExcelTemplates:

    def __init__(self, fresh=False, mode='enterprise'):
        """
            Initialization - Creates a ExcelTemplate object

            :param mode: The domain to utilize
        """
        muse = mode
        if muse.startswith('mitre-'):
            muse = mode[6:]
        if muse in ['enterprise', 'mobile']:
            self.mode = muse
            h = MatrixGen(fresh=fresh)
            self.codex = h.build_matrix(muse)
        else:
            raise BadTemplateException

    def _build_raw(self, showName=True, showID=False, subtechs=[], exclude=[]):
        """
            INTERNAL - builds a raw, not-yet-marked-up excel document based on the specifications

            :param showName: Whether or not to display names for each entry
            :param showID: Whether or not to display Technique IDs for each entry
            :param subtechs: List of all visible subtechniques
            :param exclude: List of of techniques to exclude from the matrix
            :return: a openpyxl workbook object containing the raw matrix
        """
        template, joins = MatrixGen._construct_panop(self.codex, subtechs, exclude)
        self.template = template
        wb = openpyxl.Workbook()

        sheet = wb.active

        header_template_f = Font(name='Calibri', bold=True)
        header_template_a = Alignment(horizontal='center', vertical='bottom')
        header_template_b = Border(bottom=Side(border_style='thin'))
        header_template_c = PatternFill(patternType='solid', start_color='DDDDDD', end_color='DDDDDD')

        for entry in template:
            c = sheet.cell(row=entry[0], column=entry[1])
            write_val = ''
            if showName and showID:
                write_val = self._get_ID(template[entry]) + ': ' + template[entry]
            elif showName:
                write_val = template[entry]
            elif showID:
                write_val = self._get_ID(template[entry])
            c.value = write_val
            if entry[0] == 1:
                c.font = header_template_f
                c.alignment = header_template_a
                c.border = header_template_b
                c.fill = header_template_c

        ## patch widths
        dims = {}
        sheet_handle = wb.active
        for row in sheet_handle:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet_handle.column_dimensions[col].width = value

        for marker in joins:
            sheet_handle.merge_cells(start_row=marker[0], start_column=marker[1], end_row=marker[0] + marker[2] - 1,
                                     end_column=marker[1])
            sheet_handle.merge_cells(start_row=1, start_column=marker[1], end_row=1, end_column=marker[1] + 1)
            adjust = sheet_handle.cell(row=marker[0], column=marker[1])
            adjust.alignment = Alignment(vertical='top')

        return wb

    def _get_ID(self, name):
        """
            INTERNAL - Do lookups to retrieve the ID of a technique given it's name

            :param name: The name of the technique to retrieve the ID of
            :return: The ID of the technique referenced by name
        """
        t_lookup = {'Initial Access': 'TA0001',
                    'Execution': 'TA0002',
                    'Persistence': 'TA0003',
                    'Privilege Escalation': 'TA0004',
                    'Defense Evasion': 'TA0005',
                    'Credential Access': 'TA0006',
                    'Discovery': 'TA0007',
                    'Lateral Movement': 'TA0008',
                    'Collection': 'TA0009',
                    'Command and Control': 'TA0011',
                    'Exfiltration': 'TA0010',
                    'Impact': 'TA0040'}
        for col in self.codex:
            for elm in col:
                for key in elm:
                    if key == 'tactic':
                        if elm[key] == name:
                            return t_lookup[elm[key]]
                    elif key == 'subtechs':
                        for en in elm[key]:
                            for ent in elm[key][en]:
                                for sub in ent:
                                    if name == sub:
                                        return str(ent[sub])
                    else:
                        if name == key:
                            return str(elm[key])

    def export(self, showName, showID, subtechs=[], exclude=[]):
        """
            Export a raw customized excel template


            :param showName: Whether or not to display names for each entry
            :param showID: Whether or not to display Technique IDs for each entry
            :param subtechs: List of all visible subtechniques
            :param exclude: List of of techniques to exclude from the matrix
            return: a openpyxl workbook object containing the raw matrix
        """
        return self._build_raw(showName, showID, subtechs, exclude)

    def retrieve_coords(self, techniqueID, tactic=None):
        """
            Locate the openpyxl coordinates of the provided technique for the currently loaded matrix

            :param techniqueID: The ID of the technique to locate
            :param tactic: Optional parameter to isolate the technique to a specific tactic
            :return: A tuple representing the (row, column) of the target element in the workbook
        """
        listing = []
        match = ''
        for col in self.codex:
            for elm in col:
                for name in elm:
                    if name == "subtechs":
                        for sp in elm[name]:
                            for t in elm[name][sp]:
                                for na in t:
                                    if t[na] == techniqueID:
                                        match = na
                    if elm[name] == techniqueID:
                        match = name
        for entry in self.template:
            if self.template[entry] == match:
                if tactic is not None:
                    try:
                        if self.template[(1, entry[1])] != MatrixGen.convert[tactic]:
                            continue
                    except KeyError:
                        # account for subtechniques when scanning
                        if self.template[(1, entry[1] - 1)] != MatrixGen.convert[tactic]:
                            continue
                listing.append(entry)
        return listing